"""
FantasyPoints.com Weekly NFL Statistical Projections Ingestion Adapter.
NFL +EV Betting Application (Bet365 Canada vs. Sharp Devig & FantasyPoints).
"""

from __future__ import annotations

import csv
import io
import logging
import re
from pathlib import Path
from typing import Any, Sequence

from app.adapters.base import BaseProjectionAdapter
from app.core.normalizer import PlayerNameNormalizer, TeamNormalizer
from app.schemas.projections import PlayerProjection, Position, StatCategory

logger = logging.getLogger(__name__)

# Known null tokens that convert to default 0.0
NULL_TOKENS = {
    "",
    "--",
    "-",
    "n/a",
    "na",
    "null",
    "none",
    "nan",
    "#",
    "?",
    "out",
    "ir",
    "dnp",
    "pup",
    "sus",
    "q",
    "d",
    "p",
}

# Header synonym dictionary for mapping diverse projection column headers
HEADER_SYNONYMS: dict[str, str] = {
    # Player identifiers
    "player": "player",
    "player name": "player",
    "player_name": "player",
    "name": "player",
    "athlete": "player",
    "passer": "player",
    "rusher": "player",
    "receiver": "player",
    # Team & Matchup
    "team": "team",
    "tm": "team",
    "player_team": "team",
    "franchise": "team",
    "pos": "position",
    "position": "position",
    "opp": "opponent",
    "opponent": "opponent",
    "vs": "opponent",
    "matchup": "opponent",
    # Passing Stats
    "pass att": "pass_att",
    "pass_att": "pass_att",
    "passatt": "pass_att",
    "patt": "pass_att",
    "pa": "pass_att",
    "pass attempts": "pass_att",
    "pass cmp": "pass_cmp",
    "pass_cmp": "pass_cmp",
    "passcmp": "pass_cmp",
    "pcmp": "pass_cmp",
    "cmp": "pass_cmp",
    "completions": "pass_cmp",
    "pass yds": "pass_yds",
    "pass_yds": "pass_yds",
    "passyds": "pass_yds",
    "pyds": "pass_yds",
    "pyd": "pass_yds",
    "pass yards": "pass_yds",
    "passing yds": "pass_yds",
    "passing yards": "pass_yds",
    "pass td": "pass_td",
    "pass_td": "pass_td",
    "passtd": "pass_td",
    "ptd": "pass_td",
    "pass tds": "pass_td",
    "passing td": "pass_td",
    "passing tds": "pass_td",
    "pass int": "pass_int",
    "pass_int": "pass_int",
    "passint": "pass_int",
    "pint": "pass_int",
    "int": "pass_int",
    "ints": "pass_int",
    "interceptions": "pass_int",
    "interception": "pass_int",
    # Rushing Stats
    "rush att": "rush_att",
    "rush_att": "rush_att",
    "rushatt": "rush_att",
    "ratt": "rush_att",
    "ra": "rush_att",
    "rushing att": "rush_att",
    "rushing attempts": "rush_att",
    "carries": "rush_att",
    "rush yds": "rush_yds",
    "rush_yds": "rush_yds",
    "rushyds": "rush_yds",
    "ryds": "rush_yds",
    "ryd": "rush_yds",
    "rushing yds": "rush_yds",
    "rushing yards": "rush_yds",
    "rush yards": "rush_yds",
    "rush td": "rush_td",
    "rush_td": "rush_td",
    "rushtd": "rush_td",
    "rtd": "rush_td",
    "rush tds": "rush_td",
    "rushing td": "rush_td",
    "rushing tds": "rush_td",
    # Receiving Stats
    "tgt": "targets",
    "targets": "targets",
    "target": "targets",
    "rec targets": "targets",
    "receiving targets": "targets",
    "rec": "receptions",
    "receptions": "receptions",
    "catches": "receptions",
    "reception": "receptions",
    "rec yds": "rec_yds",
    "rec_yds": "rec_yds",
    "recyds": "rec_yds",
    "reyds": "rec_yds",
    "receiving yds": "rec_yds",
    "receiving yards": "rec_yds",
    "reception yds": "rec_yds",
    "reception yards": "rec_yds",
    "rec td": "rec_td",
    "rec_td": "rec_td",
    "rectd": "rec_td",
    "rec tds": "rec_td",
    "receiving td": "rec_td",
    "receiving tds": "rec_td",
    # Anytime TD & Scoring
    "anytime td": "anytime_td",
    "anytime_td": "anytime_td",
    "anytime_td_mean": "anytime_td",
    "atd": "anytime_td",
    "td": "anytime_td",
    "tds": "anytime_td",
    "total td": "anytime_td",
    "total tds": "anytime_td",
    "touchdowns": "anytime_td",
    # Fantasy Points
    "fantasy points": "fantasy_points",
    "fantasy_points": "fantasy_points",
    "fpts": "fantasy_points",
    "fp": "fantasy_points",
    "dk pts": "fantasy_points",
    "fd pts": "fantasy_points",
    "proj": "fantasy_points",
    "points": "fantasy_points",
}


class FantasyPointsAdapter(BaseProjectionAdapter):
    """
    Production ingestion adapter for FantasyPoints.com weekly NFL player projections.
    Supports CSV, TSV, Excel (.xlsx/.xls), and pasted clipboard table text with
    delimiter auto-detection, synonym normalization, and defensive float sanitization.
    """

    adapter_id: str = "fantasypoints_ingest"

    def __init__(self, name: str = "FantasyPointsAdapter") -> None:
        super().__init__(name=name)

    @classmethod
    def sanitize_float(cls, val: Any, default: float = 0.0) -> float:
        """
        Defensively sanitize strings, floats, currency, commas, and injury annotations.
        """
        if val is None:
            return default
        if isinstance(val, (int, float)):
            return float(val)

        s = str(val).strip()
        if not s or s.lower() in NULL_TOKENS:
            return default

        # Strip injury annotations e.g. "24.5 (Q)", "18.0 [P]", "68.0 (IR)"
        s = re.sub(r"\s*[\(\[].*?[\)\]]", "", s)

        # Remove commas, currency symbols, asterisks, plus signs
        s = s.replace("$", "").replace(",", "").replace("*", "").strip()
        if s.startswith("+"):
            s = s[1:].strip()

        if not s or s.lower() in NULL_TOKENS:
            return default

        try:
            return float(s)
        except (ValueError, TypeError):
            return default

    def parse_projections(self, data: Any, **kwargs: Any) -> list[PlayerProjection]:
        """
        Parse raw projection data from file path, bytes, string, or structured object.
        """
        season = kwargs.get("season", 2026)
        week = kwargs.get("week", 1)

        if isinstance(data, (Path, str)):
            path_obj = Path(str(data))
            if path_obj.is_file():
                filename = path_obj.name.lower()
                with open(str(path_obj), "rb") as f:
                    content = f.read()
                return self.parse_file(content, filename=filename, season=season, week=week)
            else:
                return self.parse_clipboard_text(str(data), season=season, week=week)

        elif isinstance(data, bytes):
            filename = kwargs.get("filename", "upload.csv")
            return self.parse_file(data, filename=filename, season=season, week=week)

        elif hasattr(data, "to_dict") and callable(data.to_dict):
            # pandas DataFrame support
            records = data.to_dict(orient="records")
            return self._parse_dict_records(records, season=season, week=week)

        elif isinstance(data, list) and all(isinstance(item, dict) for item in data):
            return self._parse_dict_records(data, season=season, week=week)

        raise TypeError(f"Unsupported data format for FantasyPointsAdapter: {type(data)}")

    def parse_file(
        self,
        content: bytes,
        filename: str = "fantasypoints.csv",
        season: int = 2026,
        week: int = 1,
    ) -> list[PlayerProjection]:
        """
        Parse raw file bytes (CSV, TSV, or Excel .xlsx/.xls).
        """
        fn_lower = filename.lower()
        if fn_lower.endswith((".xlsx", ".xlsm", ".xltx", ".xltm", ".xls")):
            return self._parse_excel_bytes(content, season=season, week=week)

        # Standard text file parsing
        text = content.decode("utf-8", errors="replace")
        return self.parse_clipboard_text(text, season=season, week=week)

    @classmethod
    def parse_csv_text(
        cls,
        text: str,
        season: int = 2026,
        week: int = 1,
    ) -> list[PlayerProjection]:
        """
        Convenience classmethod for parsing CSV/TSV plain text.
        """
        adapter = cls()
        return adapter.parse_clipboard_text(text, season=season, week=week)

    def parse_clipboard_text(
        self,
        text: str,
        season: int = 2026,
        week: int = 1,
        inferred_pos: str | None = None,
    ) -> list[PlayerProjection]:
        """
        Parse raw pasted text from clipboard (tab, comma, pipe, or whitespace-delimited).
        """
        trimmed = text.strip()
        if not trimmed:
            return []

        lines = [l.strip() for l in trimmed.splitlines() if l.strip() and not l.strip().startswith("#")]
        if not lines:
            return []

        delimiter = self.sniff_delimiter(trimmed)

        if delimiter == r"\s{2,}":
            raw_rows = [re.split(r"\s{2,}", line) for line in lines]
        else:
            reader = csv.reader(lines, delimiter=delimiter)
            raw_rows = list(reader)

        if not raw_rows:
            return []

        header_row = raw_rows[0]
        normalized_headers = self._normalize_headers(header_row)

        dict_records: list[dict[str, Any]] = []
        for row in raw_rows[1:]:
            if not row or all(not cell.strip() for cell in row):
                continue
            row_dict: dict[str, Any] = {}
            for i, val in enumerate(row):
                if i < len(normalized_headers):
                    key = normalized_headers[i]
                    if key:
                        row_dict[key] = val.strip()
            if row_dict:
                dict_records.append(row_dict)

        return self._parse_dict_records(dict_records, season=season, week=week, inferred_pos=inferred_pos)

    def sniff_delimiter(self, text: str) -> str:
        """
        Sniff table delimiter from input text.
        """
        sample_lines = [l for l in text.splitlines() if l.strip() and not l.strip().startswith("#")][:5]
        if not sample_lines:
            return ","
        sample = "\n".join(sample_lines)

        counts = {
            "\t": sample.count("\t"),
            ",": sample.count(","),
            "|": sample.count("|"),
            ";": sample.count(";"),
        }
        best_delim = max(counts, key=counts.get)  # type: ignore[arg-type]
        if counts[best_delim] > 0:
            return best_delim

        if re.search(r"\S\s{2,}\S", sample):
            return r"\s{2,}"

        return ","

    def _normalize_headers(self, headers: Sequence[str]) -> list[str]:
        """
        Normalize column headers against HEADER_SYNONYMS.
        """
        normalized: list[str] = []
        for h in headers:
            clean = re.sub(r"[^\w\s]", "", h).strip().lower()
            clean_under = clean.replace(" ", "_")
            if clean in HEADER_SYNONYMS:
                normalized.append(HEADER_SYNONYMS[clean])
            elif clean_under in HEADER_SYNONYMS:
                normalized.append(HEADER_SYNONYMS[clean_under])
            else:
                normalized.append(clean_under)
        return normalized

    def _parse_dict_records(
        self,
        records: list[dict[str, Any]],
        season: int = 2026,
        week: int = 1,
        inferred_pos: str | None = None,
    ) -> list[PlayerProjection]:
        """
        Convert row dictionaries into normalized PlayerProjection domain schemas.
        """
        projections: list[PlayerProjection] = []

        for row in records:
            # Clean keys in row
            clean_row: dict[str, Any] = {}
            for k, v in row.items():
                if k is None:
                    continue
                k_clean = str(k).strip().lower().replace(" ", "_")
                clean_k = HEADER_SYNONYMS.get(k_clean, k_clean)
                clean_row[clean_k] = v

            raw_name = clean_row.get("player") or clean_row.get("name") or clean_row.get("player_name") or ""
            if not raw_name or str(raw_name).lower() in ("player", "total", "average", "name"):
                continue

            canonical_name = PlayerNameNormalizer.clean_name(raw_name)
            raw_team = str(clean_row.get("team") or clean_row.get("tm") or "").strip()
            team = TeamNormalizer.canonical_team(raw_team) if raw_team else "FA"

            raw_opp = str(clean_row.get("opponent") or clean_row.get("opp") or "").strip()
            opponent = TeamNormalizer.canonical_team(raw_opp) if raw_opp else None

            # Numeric extractions
            pass_att = self.sanitize_float(clean_row.get("pass_att"))
            pass_cmp = self.sanitize_float(clean_row.get("pass_cmp"))
            pass_yds = self.sanitize_float(clean_row.get("pass_yds"))
            pass_td = self.sanitize_float(clean_row.get("pass_td"))
            pass_int = self.sanitize_float(clean_row.get("pass_int"))

            rush_att = self.sanitize_float(clean_row.get("rush_att"))
            rush_yds = self.sanitize_float(clean_row.get("rush_yds"))
            rush_td = self.sanitize_float(clean_row.get("rush_td"))

            targets = self.sanitize_float(clean_row.get("targets"))
            receptions = self.sanitize_float(clean_row.get("receptions"))
            rec_yds = self.sanitize_float(clean_row.get("rec_yds"))
            rec_td = self.sanitize_float(clean_row.get("rec_td"))

            raw_atd = self.sanitize_float(clean_row.get("anytime_td"))
            fantasy_points = self.sanitize_float(clean_row.get("fantasy_points"))

            # Automatic Composite Anytime TD Calculation
            if raw_atd > 0.0:
                anytime_td_mean = raw_atd
            else:
                anytime_td_mean = round(rush_td + rec_td, 4)

            # Determine position
            pos_str = str(clean_row.get("position") or clean_row.get("pos") or "").strip().upper()
            if not pos_str or pos_str == "FLEX":
                if inferred_pos:
                    pos_str = inferred_pos.upper().strip()
                else:
                    pos_str = self._infer_position(
                        pass_att=pass_att,
                        pass_yds=pass_yds,
                        rush_att=rush_att,
                        targets=targets,
                        receptions=receptions,
                    )

            if pos_str not in Position.__members__:
                pos_str = "WR"

            stats_map: dict[StatCategory, float] = {}
            if pass_yds > 0 or pass_att > 0 or pos_str == "QB":
                stats_map[StatCategory.PASSING_YARDS] = pass_yds
                stats_map[StatCategory.PASSING_TDS] = pass_td
                stats_map[StatCategory.PASSING_INTERCEPTIONS] = pass_int
                stats_map[StatCategory.PASSING_ATTEMPTS] = pass_att
                stats_map[StatCategory.PASSING_COMPLETIONS] = pass_cmp

            if rush_yds != 0.0 or rush_att > 0 or pos_str in ("QB", "RB"):
                stats_map[StatCategory.RUSHING_YARDS] = rush_yds
                stats_map[StatCategory.RUSHING_ATTEMPTS] = rush_att
                stats_map[StatCategory.RUSHING_TDS] = rush_td

            if rec_yds > 0 or targets > 0 or receptions > 0 or pos_str in ("RB", "WR", "TE"):
                stats_map[StatCategory.RECEIVING_YARDS] = rec_yds
                stats_map[StatCategory.RECEPTIONS] = receptions
                stats_map[StatCategory.RECEIVING_TARGETS] = targets
                stats_map[StatCategory.RECEIVING_TDS] = rec_td

            if anytime_td_mean > 0.0 or pos_str in ("RB", "WR", "TE", "QB"):
                stats_map[StatCategory.ANYTIME_TD] = anytime_td_mean

            metadata_dict = {
                "raw_stats": {
                    "pass_att": pass_att,
                    "pass_cmp": pass_cmp,
                    "pass_yds": pass_yds,
                    "pass_td": pass_td,
                    "pass_int": pass_int,
                    "rush_att": rush_att,
                    "rush_yds": rush_yds,
                    "rush_td": rush_td,
                    "targets": targets,
                    "receptions": receptions,
                    "rec_yds": rec_yds,
                    "rec_td": rec_td,
                    "anytime_td_mean": anytime_td_mean,
                    "fantasy_points": fantasy_points,
                }
            }

            for stat_cat, mean_val in stats_map.items():
                proj = PlayerProjection(
                    player_name=str(raw_name).strip(),
                    canonical_name=canonical_name,
                    team=team,
                    position=pos_str,
                    opponent=opponent,
                    stat_category=stat_cat,
                    projection_mean=mean_val,
                    source=self.adapter_id,
                    season=season,
                    week=week,
                    metadata=metadata_dict,
                )
                projections.append(proj)

        return projections

    def _infer_position(
        self,
        pass_att: float,
        pass_yds: float,
        rush_att: float,
        targets: float,
        receptions: float,
    ) -> str:
        """
        Infer player position based on statistical footprint.
        """
        if pass_att >= 5.0 or pass_yds >= 40.0:
            return "QB"
        elif rush_att >= 4.0:
            return "RB"
        elif targets >= 2.0 or receptions >= 1.5:
            return "WR"
        return "WR"

    def _parse_excel_bytes(
        self,
        content: bytes,
        season: int = 2026,
        week: int = 1,
    ) -> list[PlayerProjection]:
        """
        Parse Excel workbook using openpyxl.
        """
        try:
            import openpyxl  # type: ignore[import-untyped]
        except ImportError:
            logger.warning("openpyxl not installed. Unable to parse Excel file bytes directly.")
            return []

        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        all_projections: list[PlayerProjection] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            header_row = [str(cell or "").strip() for cell in rows[0]]
            normalized_headers = self._normalize_headers(header_row)

            # Check if sheet name indicates a position
            pos_hint = sheet_name.upper().strip() if sheet_name.upper().strip() in ("QB", "RB", "WR", "TE") else None

            dict_records: list[dict[str, Any]] = []
            for row in rows[1:]:
                if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                    continue
                row_dict: dict[str, Any] = {}
                for i, val in enumerate(row):
                    if i < len(normalized_headers):
                        key = normalized_headers[i]
                        if key:
                            row_dict[key] = val
                if row_dict:
                    dict_records.append(row_dict)

            sheet_projs = self._parse_dict_records(dict_records, season=season, week=week, inferred_pos=pos_hint)
            all_projections.extend(sheet_projs)

        return all_projections


# Alias for backwards compatibility with survey specifications and tests
FantasyPointsIngestionEngine = FantasyPointsAdapter
